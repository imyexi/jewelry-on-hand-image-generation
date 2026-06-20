from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from jewelry_on_hand.models import ReferenceRow


SHEET_NAME = "分类明细"

REQUIRED_COLUMNS = (
    "序号",
    "文件名",
    "相对路径",
    "绝对路径",
    "宽度",
    "高度",
    "大小MB",
    "用途分类",
    "手链手串适用性",
    "默认使用策略",
    "风格分类",
    "场景关键词",
    "饰品类型",
    "推荐使用方式",
    "备注",
    "判断置信度",
)

REQUIRED_VALUE_COLUMNS = (
    "序号",
    "文件名",
    "相对路径",
    "绝对路径",
    "用途分类",
    "手链手串适用性",
    "默认使用策略",
    "风格分类",
    "场景关键词",
    "饰品类型",
    "推荐使用方式",
    "判断置信度",
)


def load_reference_rows(workbook_path: str | Path) -> list[ReferenceRow]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        return _load_rows_from_workbook(workbook)
    finally:
        workbook.close()


def _load_rows_from_workbook(workbook: Any) -> list[ReferenceRow]:
    if SHEET_NAME not in workbook.sheetnames:
        sheet_names = "、".join(str(name) for name in workbook.sheetnames)
        raise ValueError(f"工作簿缺少工作表：{SHEET_NAME}；实际工作表：{sheet_names}")

    worksheet = workbook[SHEET_NAME]
    rows_iter = worksheet.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration as exc:
        raise ValueError(f"工作表 {SHEET_NAME} 缺少表头") from exc

    column_index = {name: index for index, name in enumerate(header) if name is not None}
    missing_columns = [name for name in REQUIRED_COLUMNS if name not in column_index]
    if missing_columns:
        missing_text = "、".join(missing_columns)
        actual_headers = "、".join(str(name) for name in header if name is not None)
        raise ValueError(
            f"工作表 {SHEET_NAME} 缺少必需列：{missing_text}；实际表头：{actual_headers}"
        )

    reference_rows: list[ReferenceRow] = []
    for row_number, values in enumerate(rows_iter, start=2):
        if _is_blank_row(values):
            continue
        row_data = _row_data(values, column_index, row_number)
        absolute_path = row_data["absolute_path"]
        reference_rows.append(
            ReferenceRow(
                index=row_data["index"],
                file_name=row_data["file_name"],
                relative_path=row_data["relative_path"],
                absolute_path=absolute_path,
                width=row_data["width"],
                height=row_data["height"],
                size_mb=row_data["size_mb"],
                purpose_category=row_data["purpose_category"],
                bracelet_applicability=row_data["bracelet_applicability"],
                default_strategy=row_data["default_strategy"],
                style_category=row_data["style_category"],
                scene_keywords=row_data["scene_keywords"],
                jewelry_type=row_data["jewelry_type"],
                recommended_usage=row_data["recommended_usage"],
                notes=row_data["notes"],
                confidence=row_data["confidence"],
                file_exists=absolute_path.is_file(),
            )
        )
    return reference_rows


def _row_data(values: tuple[Any, ...], column_index: dict[str, int], row_number: int) -> dict[str, Any]:
    for column_name in REQUIRED_VALUE_COLUMNS:
        _required_cell(values, column_index, row_number, column_name)

    absolute_path_value = _required_cell(values, column_index, row_number, "绝对路径")
    if not isinstance(absolute_path_value, (str, Path)):
        raise _cell_error(row_number, "绝对路径", "必须是路径字符串")

    return {
        "index": _required_cell(values, column_index, row_number, "序号"),
        "file_name": _required_cell(values, column_index, row_number, "文件名"),
        "relative_path": _required_cell(values, column_index, row_number, "相对路径"),
        "absolute_path": Path(absolute_path_value),
        "width": _optional_number_cell(values, column_index, "宽度"),
        "height": _optional_number_cell(values, column_index, "高度"),
        "size_mb": _optional_number_cell(values, column_index, "大小MB"),
        "purpose_category": _required_cell(values, column_index, row_number, "用途分类"),
        "bracelet_applicability": _required_cell(
            values, column_index, row_number, "手链手串适用性"
        ),
        "default_strategy": _required_cell(values, column_index, row_number, "默认使用策略"),
        "style_category": _required_cell(values, column_index, row_number, "风格分类"),
        "scene_keywords": _required_cell(values, column_index, row_number, "场景关键词"),
        "jewelry_type": _required_cell(values, column_index, row_number, "饰品类型"),
        "recommended_usage": _required_cell(values, column_index, row_number, "推荐使用方式"),
        "notes": _optional_text_cell(values, column_index, "备注"),
        "confidence": _required_cell(values, column_index, row_number, "判断置信度"),
    }


def _required_cell(
    values: tuple[Any, ...],
    column_index: dict[str, int],
    row_number: int,
    column_name: str,
) -> Any:
    value = _cell_value(values, column_index[column_name])
    if value is None or (isinstance(value, str) and not value.strip()):
        raise _cell_error(row_number, column_name, "不能为空")
    if isinstance(value, str):
        return value.strip()
    return value


def _optional_number_cell(values: tuple[Any, ...], column_index: dict[str, int], column_name: str) -> Any:
    value = _cell_value(values, column_index[column_name])
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    return value


def _optional_text_cell(values: tuple[Any, ...], column_index: dict[str, int], column_name: str) -> str:
    value = _cell_value(values, column_index[column_name])
    if value is None or (isinstance(value, str) and not value.strip()):
        return ""
    if not isinstance(value, str):
        raise _cell_error(0, column_name, "必须是字符串")
    return value


def _cell_value(values: tuple[Any, ...], index: int) -> Any:
    if index >= len(values):
        return None
    return values[index]


def _is_blank_row(values: tuple[Any, ...]) -> bool:
    return all(value is None or (isinstance(value, str) and not value.strip()) for value in values)


def _cell_error(row_number: int, column_name: str, message: str) -> ValueError:
    row_text = f"第 {row_number} 行" if row_number else "数据行"
    return ValueError(f"工作表 {SHEET_NAME} {row_text} 列「{column_name}」{message}")
