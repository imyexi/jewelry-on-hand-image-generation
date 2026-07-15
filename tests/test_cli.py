import json
from hashlib import sha256
from pathlib import Path

import pytest
from openpyxl import Workbook

from jewelry_on_hand.models import (
    ProductAnalysis,
    ProductFidelityConstraints,
    ReferenceRow,
    ScoredReference,
)
from jewelry_on_hand.product_fidelity import build_product_fidelity_constraints
from jewelry_on_hand.product_analysis import load_product_analysis
from jewelry_on_hand.product_analysis import product_analysis_to_dict
from jewelry_on_hand.qc import build_qc_checklist, qc_check_id
from jewelry_on_hand.qc_review import (
    build_reference_preservation_checklist,
    write_qc_review_page,
)
from jewelry_on_hand.reference_composition import (
    ReferenceCompositionSnapshot,
    ReferencePose,
    ReplacementTarget,
    build_candidate_snapshot,
    reference_composition_sha256,
)
from jewelry_on_hand.review_decision import write_review_bundle
from jewelry_on_hand.review_package import write_review_package
from jewelry_on_hand.run_paths import RunPaths, read_json, write_json


@pytest.mark.parametrize("invalid_notes", [["数组"], {"对象": True}, True, 1])
def test_qc_cli_rejects_non_string_fidelity_check_notes_without_writing(
    tmp_path, capsys, invalid_notes
):
    from jewelry_on_hand.cli import main

    generation_dir, checks_json, checklist_json, reference_json = (
        _ready_cli_qc_generation(tmp_path)
    )
    write_json(
        checks_json,
        [
            {
                "name": "主吊坠",
                "question": "主吊坠是否保持原连接？",
                "result": "rerun",
                "notes": invalid_notes,
            }
        ],
    )

    result = main(
        [
            "qc",
            "--generation-dir",
            str(generation_dir),
            "--status",
            "rerun",
            "--failed",
            "需要复核",
            "--fidelity-checks-json",
            str(checks_json),
            "--checklist-checks-json",
            str(checklist_json),
            "--reference-preservation-checks-json",
            str(reference_json),
        ]
    )

    assert result == 1
    assert "fidelity_checks.notes 必须是字符串或 null" in capsys.readouterr().err
    assert not (generation_dir / "qc.json").exists()


def test_qc_cli_requires_all_three_check_json_arguments(tmp_path):
    from jewelry_on_hand.cli import main

    with pytest.raises(SystemExit):
        main(
            [
                "qc",
                "--generation-dir",
                str(tmp_path / "generation" / "01"),
                "--status",
                "pass",
            ]
        )


def test_qc_cli_writes_complete_modern_three_layer_result(tmp_path):
    from jewelry_on_hand.cli import main

    generation_dir, fidelity, checklist, reference = _ready_cli_qc_generation(
        tmp_path
    )

    result = main(
        [
            "qc",
            "--generation-dir",
            str(generation_dir),
            "--status",
            "pass",
            "--passed",
            "三层人工复核完成",
            "--notes",
            "逐项对照四栏页面",
            "--fidelity-checks-json",
            str(fidelity),
            "--checklist-checks-json",
            str(checklist),
            "--reference-preservation-checks-json",
            str(reference),
        ]
    )

    payload = read_json(generation_dir / "qc.json")
    assert result == 0
    assert payload["status"] == "pass"
    assert payload["reference_preservation_checks"]
    assert "fidelity_checks" in payload
    assert payload["checklist_checks"]


def test_qc_cli_rejects_legacy_or_hero_generation_without_writing(tmp_path, capsys):
    from jewelry_on_hand.cli import main

    generation_dir, fidelity, checklist, reference = _ready_cli_qc_generation(
        tmp_path
    )
    (generation_dir / "input-manifest.json").unlink()
    args = [
        "qc",
        "--generation-dir",
        str(generation_dir),
        "--status",
        "pass",
        "--fidelity-checks-json",
        str(fidelity),
        "--checklist-checks-json",
        str(checklist),
        "--reference-preservation-checks-json",
        str(reference),
    ]

    assert main(args) == 1
    assert "历史离线 QC 仅可只读" in capsys.readouterr().err
    assert not (generation_dir / "qc.json").exists()

    write_json(
        generation_dir / "input-manifest.json",
        {"schema_version": 1, "output_role": "hero"},
    )
    assert main(args) == 1
    assert "主图" in capsys.readouterr().err
    assert not (generation_dir / "qc.json").exists()


def test_reference_ensure_fields_help_and_output_are_chinese(monkeypatch, capsys):
    from jewelry_on_hand.cli import main

    with pytest.raises(SystemExit) as exc_info:
        main(["reference-ensure-fields", "--help"])
    assert exc_info.value.code == 0
    help_text = capsys.readouterr().out
    assert "补齐飞书参考图库缺失的 AI 字段" in help_text
    assert "???" not in help_text

    monkeypatch.setattr(
        "jewelry_on_hand.cli.ensure_enrichment_fields",
        lambda _config: ["适用产品类型", "适用展示模式"],
    )
    assert main(["reference-ensure-fields"]) == 0
    created_output = capsys.readouterr().out
    assert "已创建 AI 补齐字段：适用产品类型、适用展示模式" in created_output

    monkeypatch.setattr(
        "jewelry_on_hand.cli.ensure_enrichment_fields", lambda _config: []
    )
    assert main(["reference-ensure-fields"]) == 0
    assert "飞书参考图库已包含全部 AI 补齐字段" in capsys.readouterr().out


def test_prepare_review_help_limits_product_detail_image_to_review_context(capsys):
    from jewelry_on_hand.cli import main

    with pytest.raises(SystemExit) as exc_info:
        main(["prepare-review", "--help"])

    assert exc_info.value.code == 0
    help_text = capsys.readouterr().out
    outdated_help_text = "作为审核和" + "生成的产品身份图"
    assert outdated_help_text not in help_text
    normalized_help_text = " ".join(help_text.split())
    assert "仅用于 review、结构分析、canonical 约束和人工 QC" in normalized_help_text
    assert "不进入模型" in help_text


def make_catalog(path, ref):
    wb = Workbook()
    ws = wb.active
    ws.title = "分类明细"
    ws.append(
        [
            "序号",
            "文件名",
            "相对路径",
            "绝对路径",
            "宽度",
            "高度",
            "大小MB",
            "用途分类",
            "手链手串适用性",
            "默认使用策略",
            "风格分类",
            "场景关键词",
            "饰品类型",
            "推荐使用方式",
            "备注",
            "判断置信度",
        ]
    )
    for i in range(1, 4):
        ws.append(
            [
                i,
                f"ref{i}.jpg",
                f"ref{i}.jpg",
                str(ref),
                100,
                200,
                0.1,
                "上手姿势/手模构图参考",
                "是：可用于手链/手串",
                "常规可优先使用",
                "暗调闪光",
                "车内 闪光",
                "手链/手串",
                "近景手腕",
                "手腕/前臂露出面积足",
                "高",
            ]
        )
        ws.cell(row=ws.max_row, column=8, value="手部佩戴图")
    wb.save(path)


def make_analysis(path):
    data = {
        "product_type": "手链/手串",
        "wear_position": "手腕",
        "visible_appearance": "深红主珠",
        "color_family": ["深红"],
        "style_mood": "暗调闪光",
        "composition": "手腕近景",
        "product_dimensions": {},
        "needs_full_front_display": True,
        "special_requirements": ["保留主珠"],
    }
    path.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    return data


def make_modern_analysis(path, **overrides):
    data = {
        "product_type": "疑似项链",
        "detected_product_type": "unknown",
        "confirmed_product_type": "unknown",
        "classification_confidence": "low",
        "classification_evidence": ["颈部结构被遮挡"],
        "classification_source": "auto_uncertain",
        "display_mode": "worn",
        "source_image_type": "worn_source",
        "wear_position": "颈部和锁骨",
        "visible_appearance": "双层珠链，中央结构不清晰",
        "color_family": ["白色"],
        "style_mood": "精致",
        "composition": "胸前近景",
        "product_dimensions": {},
        "needs_full_front_display": True,
        "special_requirements": [],
        "layer_count": 1,
        "length_category": None,
        "chain_or_strand_type": "beaded",
        "has_pendant": False,
        "pendant_count": 0,
        "pendant_layer": None,
        "pendant_position": None,
        "pendant_orientation": None,
        "connection_structure": None,
        "symmetry": None,
        "occluded_parts": ["后颈扣头"],
        "uncertain_details": ["中央结构"],
        "is_independent_multi_item": False,
    }
    data.update(overrides)
    write_json(path, data)
    return data


def declare_scene_output_role(run_root, role="hand_worn"):
    write_json(
        Path(run_root) / "analysis" / "output_role.json",
        {"output_role": role},
    )


def prepare_snapshot_decision_run(run_root, *, role="hand_worn"):
    paths = RunPaths.create(Path(run_root).parent, Path(run_root).name)
    (paths.input_dir / "product-on-hand.jpg").write_bytes(b"product")
    reference = paths.input_dir / "decision-reference.jpg"
    reference.write_bytes(b"decision-reference")
    analysis_path = paths.analysis_dir / "product_analysis.json"
    analysis_data = (
        read_json(analysis_path)
        if analysis_path.is_file()
        else make_analysis(analysis_path)
    )
    constraints = build_product_fidelity_constraints(
        ProductAnalysis.from_dict(analysis_data)
    ).to_dict()
    if constraints["review_status"] == "pending":
        constraints["review_status"] = "confirmed"
    write_json(paths.analysis_dir / "product_fidelity_constraints.json", constraints)
    declare_scene_output_role(run_root, role)
    scored = ScoredReference(
        ReferenceRow(
            index=1,
            file_name=reference.name,
            relative_path=reference.name,
            absolute_path=reference,
            width=100,
            height=200,
            size_mb=0.1,
            purpose_category="手部佩戴图",
            bracelet_applicability="是",
            default_strategy="常规可优先使用",
            style_category="暗调闪光",
            scene_keywords="车内",
            jewelry_type="手链/手串",
            recommended_usage="近景手腕",
            notes="正面视角，主体居中，无文字或 UI",
            confidence="高",
            file_exists=True,
            framing="手部近景",
            visible_body_regions="左手腕 / 前臂完整露出",
            product_visibility="展示面积充足，大于 35%",
            collar_type="无衣领",
            clothing_occlusion_risk="衣物无遮挡",
            pose_keywords="身体未入镜，前臂自然抬起",
            existing_jewelry="左手腕原有手链",
            crop_risk="裁切风险低",
            hand_side="左手",
            hand_orientation="手背朝向镜头",
        ),
        99,
        1,
        ("匹配",),
        (),
        ("原有手链",),
    )
    snapshot = build_candidate_snapshot(
        load_product_analysis(paths.analysis_dir / "product_analysis.json"),
        scored,
        role,
    )
    review_copy = paths.review_dir / f"rank-1-{reference.name}"
    review_copy.write_bytes(reference.read_bytes())
    digest = sha256(reference.read_bytes()).hexdigest()
    selected = scored.to_dict()
    selected["selected_reference"] = str(review_copy.resolve())
    selected["source_sha256"] = digest
    selected["review_sha256"] = digest
    selected["metadata"]["source_sha256"] = digest
    selected["metadata"]["review_sha256"] = digest
    write_json(paths.analysis_dir / "selected_references.json", [selected])
    write_json(
        paths.analysis_dir / "reference_composition_snapshots.json",
        [snapshot.to_dict()],
    )
    return snapshot


def test_generate_cli_loads_snapshot_analysis_and_canonical_for_input_manifest(
    tmp_path,
    monkeypatch,
):
    from jewelry_on_hand.cli import main

    run_root = tmp_path / "runs" / "demo"
    snapshot = prepare_snapshot_decision_run(run_root)
    paths = RunPaths(root=run_root)
    write_review_bundle(
        paths,
        {
            "action": "generate_rank_1",
            "selected_ranks": [1],
            "fidelity_confirmed": True,
            "output_role": "hand_worn",
        },
    )
    prompt_snapshots = []
    generation_calls = []

    def fake_build_prompt(
        product,
        reference,
        fidelity_constraints,
        output_role,
        reference_snapshot,
    ):
        prompt_snapshots.append(reference_snapshot)
        return "prompt-rank-1"

    def fake_run_generation(
        paths,
        product_image,
        prompts_by_rank,
        helper_script,
        wait=True,
        *,
        reference_snapshot,
        product_analysis_path,
        fidelity_constraints_path,
    ):
        generation_calls.append(
            {
                "paths": paths,
                "product_image": product_image,
                "prompts": prompts_by_rank,
                "wait": wait,
                "snapshot": reference_snapshot,
                "analysis": product_analysis_path,
                "canonical": fidelity_constraints_path,
            }
        )
        return [paths.generation_dir / "01"]

    monkeypatch.setattr("jewelry_on_hand.cli.build_prompt", fake_build_prompt)
    monkeypatch.setattr("jewelry_on_hand.cli.run_generation", fake_run_generation)

    assert main(
        [
            "generate",
            "--run-root",
            str(run_root),
            "--helper-script",
            str(tmp_path / "helper.py"),
            "--no-wait",
        ]
    ) == 0
    assert prompt_snapshots == [snapshot]
    assert len(generation_calls) == 1
    call = generation_calls[0]
    assert call["snapshot"] == snapshot
    assert call["analysis"] == paths.analysis_dir / "product_analysis.json"
    assert call["canonical"] == paths.analysis_dir / "product_fidelity_constraints.json"
    assert call["prompts"] == {1: "prompt-rank-1"}
    assert call["wait"] is False


def make_constraints(
    path,
    review_status="confirmed",
    must_keep=None,
    analysis_data=None,
):
    must_keep_was_default = must_keep is None
    if must_keep is None:
        must_keep = [
            {
                "name": "白水晶随形",
                "source_text": "白水晶随形",
                "normalized_keyword": "随形",
                "location": "主珠右侧",
                "visual_shape": "透明不规则随形，非圆珠",
                "relationship": "位于两颗圆珠之间",
                "forbid": ["改成圆珠"],
                "qc_question": "白水晶随形是否仍是不规则透明异形珠",
            }
        ]
    data = {
        "schema_version": 1,
        "source": {
            "product_image": "input/product-on-hand.jpg",
            "product_analysis": "analysis/product_analysis.json",
        },
        "detected_keywords": ["随形"] if must_keep else [],
        "must_keep": must_keep,
        "must_not_change": ["珠子排列顺序"],
        "needs_user_review": bool(must_keep),
        "detail_crop_recommended": bool(must_keep),
        "review_status": review_status,
    }
    analysis_path = Path(path).parent / "product_analysis.json"
    if analysis_data is None and analysis_path.is_file():
        analysis_data = read_json(analysis_path)
    if analysis_data is not None:
        try:
            analysis = ProductAnalysis.from_dict(analysis_data)
            built = build_product_fidelity_constraints(analysis)
        except ValueError:
            built = None
        if built is not None:
            data["schema_version"] = built.schema_version
            data["source"]["product_analysis_sha256"] = built.source[
                "product_analysis_sha256"
            ]
            data["source"]["product_type"] = built.source["product_type"]
            data["must_not_change"] = list(built.must_not_change)
            if built.pendant_semantics is not None:
                data["pendant_semantics"] = built.pendant_semantics.to_dict()
            if must_keep_was_default and analysis.normalized_product_type.value != "bracelet":
                data["must_keep"] = [item.to_dict() for item in built.must_keep]
                data["detected_keywords"] = list(built.detected_keywords)
                data["needs_user_review"] = built.needs_user_review
                data["detail_crop_recommended"] = built.detail_crop_recommended
    write_json(path, data)
    return data


def test_prepare_review_cli_creates_review_html(tmp_path, monkeypatch):
    from jewelry_on_hand.cli import main

    product = tmp_path / "product.jpg"
    product.write_bytes(b"product")
    ref = tmp_path / "ref.jpg"
    ref.write_bytes(b"ref")
    catalog = tmp_path / "catalog.xlsx"
    make_catalog(catalog, ref)
    analysis = tmp_path / "analysis.json"
    make_analysis(analysis)
    monkeypatch.setattr(
        "jewelry_on_hand.cli.sync_and_load_reference_rows",
        lambda config: __import__("jewelry_on_hand.reference_catalog", fromlist=["load_reference_rows"]).load_reference_rows(catalog),
    )

    assert (
        main(
            [
                "prepare-review",
                "--product-image",
                str(product),
                "--analysis-json",
                str(analysis),
                "--output-root",
                str(tmp_path / "runs"),
                    "--run-id",
                    "demo",
                    "--output-role",
                    "hand_worn",
            ]
        )
        == 0
    )

    run_root = tmp_path / "runs" / "demo"
    assert (run_root / "review" / "review.html").is_file()
    assert (run_root / "analysis" / "product_analysis_prompt.txt").is_file()
    assert (run_root / "analysis" / "product_fidelity_constraints.json").is_file()
    fidelity = read_json(run_root / "analysis" / "product_fidelity_constraints.json")
    assert fidelity["schema_version"] == 1
    assert "must_keep" in fidelity
    assert fidelity["review_status"] in {"pending", "not_applicable"}
    assert read_json(run_root / "analysis" / "product_analysis.json")["product_type"] == "手链/手串"
    assert not (run_root / "review" / "review_decision.json").exists()


def test_prepare_review_cli_rejects_existing_non_empty_run_without_overwrite(tmp_path, monkeypatch):
    from jewelry_on_hand.cli import main

    run_root = tmp_path / "runs" / "demo"
    (run_root / "input").mkdir(parents=True)
    (run_root / "review").mkdir()
    old_product = run_root / "input" / "product-on-hand.jpg"
    old_product.write_bytes(b"old-product")
    write_json(run_root / "review" / "review_decision.json", {"action": "generate_rank_1", "selected_ranks": [1]})

    product = tmp_path / "new-product.jpg"
    product.write_bytes(b"new-product")
    ref = tmp_path / "ref.jpg"
    ref.write_bytes(b"ref")
    catalog = tmp_path / "catalog.xlsx"
    make_catalog(catalog, ref)
    analysis = tmp_path / "analysis.json"
    make_analysis(analysis)
    monkeypatch.setattr(
        "jewelry_on_hand.cli.sync_and_load_reference_rows",
        lambda config: __import__("jewelry_on_hand.reference_catalog", fromlist=["load_reference_rows"]).load_reference_rows(catalog),
    )

    assert (
        main(
            [
                "prepare-review",
                "--product-image",
                str(product),
                "--analysis-json",
                str(analysis),
                "--output-root",
                str(tmp_path / "runs"),
                "--run-id",
                "demo",
            ]
        )
        != 0
    )

    assert old_product.read_bytes() == b"old-product"
    assert (run_root / "review" / "review_decision.json").is_file()
    assert not (run_root / "review" / "review.html").exists()


def test_prepare_review_cli_writes_dimensions_json_and_includes_it_in_prompt(tmp_path, monkeypatch):
    from jewelry_on_hand.cli import main

    product = tmp_path / "product.jpg"
    product.write_bytes(b"product")
    ref = tmp_path / "ref.jpg"
    ref.write_bytes(b"ref")
    catalog = tmp_path / "catalog.xlsx"
    make_catalog(catalog, ref)
    analysis = tmp_path / "analysis.json"
    make_analysis(analysis)
    monkeypatch.setattr(
        "jewelry_on_hand.cli.sync_and_load_reference_rows",
        lambda config: __import__("jewelry_on_hand.reference_catalog", fromlist=["load_reference_rows"]).load_reference_rows(catalog),
    )
    dimensions = tmp_path / "dimensions.json"
    dimensions.write_text(
        json.dumps(
            {"bead_diameter_mm": 12.5, "dimension_source": "用户提供尺寸"},
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    assert (
        main(
            [
                "prepare-review",
                "--product-image",
                str(product),
                "--analysis-json",
                str(analysis),
                "--output-root",
                str(tmp_path / "runs"),
                "--run-id",
                "demo",
                    "--dimensions-json",
                    str(dimensions),
                    "--output-role",
                    "hand_worn",
            ]
        )
        == 0
    )

    run_root = tmp_path / "runs" / "demo"
    assert read_json(run_root / "input" / "product_dimensions.json") == {
        "bead_diameter_mm": 12.5,
        "dimension_source": "用户提供尺寸",
    }
    prompt = (run_root / "analysis" / "product_analysis_prompt.txt").read_text(encoding="utf-8")
    assert "12.5" in prompt
    assert "用户提供尺寸" in prompt


def test_prepare_review_cli_without_analysis_json_writes_prompt_only(tmp_path):
    from jewelry_on_hand.cli import main

    product = tmp_path / "product.jpg"
    product.write_bytes(b"product")
    ref = tmp_path / "ref.jpg"
    ref.write_bytes(b"ref")
    catalog = tmp_path / "catalog.xlsx"
    make_catalog(catalog, ref)

    assert (
        main(
            [
                "prepare-review",
                "--product-image",
                str(product),
                "--output-root",
                str(tmp_path / "runs"),
                    "--run-id",
                    "demo",
                    "--output-role",
                    "hand_worn",
            ]
        )
        != 0
    )

    run_root = tmp_path / "runs" / "demo"
    assert (run_root / "analysis" / "product_analysis_prompt.txt").is_file()
    assert not (run_root / "analysis" / "product_analysis.json").exists()
    assert not (run_root / "review" / "review.html").exists()
    assert not (run_root / "review" / "review_decision.json").exists()


def test_record_decision_cli_writes_and_normalizes_generate_rank_1(tmp_path):
    from jewelry_on_hand.cli import main

    run_root = tmp_path / "runs" / "demo"
    prepare_snapshot_decision_run(run_root)

    assert (
        main(
            [
                "record-decision",
                "--run-root",
                str(run_root),
                "--action",
                "generate_rank_1",
                "--fidelity-confirmed",
                "--output-role",
                "hand_worn",
            ]
        )
        == 0
    )

    decision = read_json(run_root / "review" / "review_decision.json")
    assert decision["action"] == "generate_rank_1"
    assert decision["selected_ranks"] == [1]
    assert decision["fidelity_confirmed"] is True
    assert decision["fidelity_constraints_path"] == (
        "analysis/product_fidelity_constraints.json"
    )
    assert decision["output_role"] == "hand_worn"
    assert len(decision["reference_snapshot_sha256"]) == 64


def test_record_decision_cli_returns_nonzero_for_invalid_selected_rank(tmp_path):
    from jewelry_on_hand.cli import main

    run_root = tmp_path / "runs" / "demo"

    assert (
        main(
            [
                "record-decision",
                "--run-root",
                str(run_root),
                "--action",
                "generate_selected",
                "--selected-ranks",
                "abc",
            ]
        )
        != 0
    )

    assert not (run_root / "review" / "review_decision.json").exists()


def test_record_decision_cli_rejects_generate_rank_1_with_rank_2(tmp_path):
    from jewelry_on_hand.cli import main

    run_root = tmp_path / "runs" / "demo"

    assert (
        main(
            [
                "record-decision",
                "--run-root",
                str(run_root),
                "--action",
                "generate_rank_1",
                "--selected-ranks",
                "2",
                "--fidelity-confirmed",
            ]
        )
        != 0
    )

    assert not (run_root / "review" / "review_decision.json").exists()


def test_record_decision_cli_rejects_duplicate_selected_ranks(tmp_path):
    from jewelry_on_hand.cli import main

    run_root = tmp_path / "runs" / "demo"

    assert (
        main(
            [
                "record-decision",
                "--run-root",
                str(run_root),
                "--action",
                "generate_multiple",
                "--selected-ranks",
                "1",
                "--selected-ranks",
                "1",
                "--fidelity-confirmed",
            ]
        )
        != 0
    )

    assert not (run_root / "review" / "review_decision.json").exists()


def test_prepare_review_cli_rejects_hero_before_creating_run(tmp_path):
    from jewelry_on_hand.cli import main
    from jewelry_on_hand.output_roles import OutputRole

    product = tmp_path / "product.jpg"
    product.write_bytes(b"product")
    analysis = tmp_path / "analysis.json"
    make_analysis(analysis)
    output_root = tmp_path / "runs"

    assert OutputRole.HERO.value == "hero"
    assert main([
        "prepare-review", "--product-image", str(product), "--analysis-json", str(analysis),
        "--output-root", str(output_root), "--run-id", "hero", "--output-role", "hero",
    ]) != 0

    assert not (output_root / "hero").exists()


@pytest.mark.parametrize("run_role", ["hand_worn", "lifestyle"])
def test_record_decision_cli_rejects_explicit_hero_for_valid_run_without_writing_decision(
    tmp_path,
    capsys,
    run_role,
):
    from jewelry_on_hand.cli import main

    run_root = tmp_path / "runs" / run_role
    write_json(
        run_root / "analysis" / "output_role.json",
        {"output_role": run_role},
    )

    assert main([
        "record-decision", "--run-root", str(run_root), "--action", "rerank",
        "--output-role", "hero",
    ]) != 0

    assert "主图 Skill" in capsys.readouterr().err
    assert not (run_root / "review" / "review_decision.json").exists()


def test_generate_cli_rejects_hero_before_calling_generation_helper(
    tmp_path, monkeypatch, capsys
):
    from jewelry_on_hand.cli import main

    run_root = tmp_path / "runs" / "hero"
    analysis_dir = run_root / "analysis"
    review_dir = run_root / "review"
    input_dir = run_root / "input"
    analysis_dir.mkdir(parents=True)
    review_dir.mkdir()
    input_dir.mkdir()
    product = input_dir / "product-on-hand.jpg"
    product.write_bytes(b"product")
    reference = tmp_path / "reference.jpg"
    reference.write_bytes(b"reference")
    make_analysis(analysis_dir / "product_analysis.json")
    make_constraints(analysis_dir / "product_fidelity_constraints.json")
    write_json(
        analysis_dir / "selected_references.json",
        [_selected_reference_payload(1, reference)],
    )
    write_json(analysis_dir / "output_role.json", {"output_role": "hero"})
    write_json(
        review_dir / "review_decision.json",
        {
            "action": "generate_rank_1",
            "selected_ranks": [1],
            "fidelity_confirmed": True,
            "output_role": "hero",
        },
    )

    helper_called = False

    def fail_if_called(*_args, **_kwargs):
        nonlocal helper_called
        helper_called = True
        raise AssertionError("生成 helper 不得被调用")

    monkeypatch.setattr("jewelry_on_hand.cli.run_generation", fail_if_called)

    assert main([
        "generate", "--run-root", str(run_root),
        "--helper-script", str(tmp_path / "helper.py"),
    ]) != 0

    assert "主图 Skill" in capsys.readouterr().err
    assert helper_called is False


def test_record_decision_cli_rejects_full_late_necklace_correction(tmp_path, capsys):
    from jewelry_on_hand.cli import main

    run_root = tmp_path / "runs" / "demo"
    analysis_path = run_root / "analysis" / "product_analysis.json"
    original = make_modern_analysis(analysis_path)
    make_constraints(
        run_root / "analysis" / "product_fidelity_constraints.json",
        must_keep=[],
    )
    corrected_analysis = original | {
        "confirmed_product_type": "pendant_necklace",
        "classification_source": "manual_override",
        "source_image_type": "worn_source",
        "display_mode": "hand_held",
        "layer_count": 2,
        "length_category": "collarbone",
        "has_pendant": True,
        "pendant_count": 1,
        "pendant_layer": 2,
        "pendant_position": "front_center",
        "pendant_orientation": "front_facing",
        "connection_structure": "metal_bail",
        "is_independent_multi_item": False,
    }
    imported_constraints = build_product_fidelity_constraints(
        ProductAnalysis.from_dict(corrected_analysis)
    ).to_dict()
    imported_path = tmp_path / "corrected-pendant-constraints.json"
    write_json(imported_path, imported_constraints)
    declare_scene_output_role(run_root)

    assert main(
        [
            "record-decision",
            "--run-root",
            str(run_root),
            "--action",
            "generate_rank_1",
            "--fidelity-confirmed",
            "--fidelity-constraints-path",
            str(imported_path),
            "--confirmed-product-type",
            "pendant_necklace",
            "--source-image-type",
            "worn_source",
            "--display-mode",
            "hand_held",
            "--layer-count",
            "2",
            "--length-category",
            "collarbone",
            "--has-pendant",
            "--pendant-count",
            "1",
            "--pendant-layer",
            "2",
            "--pendant-position",
            "front_center",
            "--pendant-orientation",
            "front_facing",
            "--connection-structure",
            "metal_bail",
            "--no-independent-multi-item",
            "--output-role",
            "hand_worn",
        ]
    ) != 0

    assert "重新执行 prepare-review" in capsys.readouterr().err
    assert read_json(analysis_path) == original
    assert not (run_root / "review" / "review_decision.json").exists()


def test_record_decision_cli_rejects_partial_late_necklace_correction(tmp_path, capsys):
    from jewelry_on_hand.cli import main

    run_root = tmp_path / "runs" / "demo"
    analysis_path = run_root / "analysis" / "product_analysis.json"
    original = make_modern_analysis(
        analysis_path,
        product_type="普通项链",
        detected_product_type="necklace",
        confirmed_product_type="necklace",
        classification_confidence="high",
        classification_evidence=["完整链条围绕颈部"],
        classification_source="auto_confirmed",
        layer_count=2,
        length_category="upper_chest",
    )
    make_constraints(
        run_root / "analysis" / "product_fidelity_constraints.json",
        must_keep=[],
    )
    declare_scene_output_role(run_root)

    assert main(
        [
            "record-decision",
            "--run-root",
            str(run_root),
            "--action",
            "generate_rank_1",
            "--fidelity-confirmed",
            "--display-mode",
            "hand_held",
            "--output-role",
            "hand_worn",
        ]
    ) != 0

    assert "重新执行 prepare-review" in capsys.readouterr().err
    assert read_json(analysis_path) == original
    assert not (run_root / "review" / "review_decision.json").exists()


def test_record_decision_cli_rejects_illegal_manual_correction_without_writing(tmp_path, capsys):
    from jewelry_on_hand.cli import main

    run_root = tmp_path / "runs" / "demo"
    analysis_path = run_root / "analysis" / "product_analysis.json"
    original = make_modern_analysis(
        analysis_path,
        product_type="普通项链",
        detected_product_type="necklace",
        confirmed_product_type="necklace",
        length_category="collarbone",
    )
    declare_scene_output_role(run_root)

    assert main(
        [
            "record-decision",
            "--run-root",
            str(run_root),
            "--action",
            "generate_rank_1",
            "--fidelity-confirmed",
            "--source-image-type",
            "flat_lay_source",
            "--output-role",
            "hand_worn",
        ]
    ) != 0

    assert "重新执行 prepare-review" in capsys.readouterr().err
    assert read_json(analysis_path) == original
    assert not (run_root / "review" / "review_decision.json").exists()


def test_record_decision_cli_rejects_unknown_and_pendant_only_categories(tmp_path, capsys):
    from jewelry_on_hand.cli import main

    for index, (product_type, expected) in enumerate(
        (
            ("unknown", "必须先人工纠正"),
            ("pendant_only", "禁止自动补链"),
        )
    ):
        run_root = tmp_path / "runs" / f"demo-{index}"
        make_modern_analysis(run_root / "analysis" / "product_analysis.json")
        declare_scene_output_role(run_root)

        assert main(
            [
                "record-decision",
                "--run-root",
                str(run_root),
                "--action",
                "generate_rank_1",
                "--fidelity-confirmed",
                "--output-role",
                "hand_worn",
                "--output-role",
                "hand_worn",
                "--confirmed-product-type",
                product_type,
                "--output-role",
                "hand_worn",
            ]
        ) != 0
        assert expected in capsys.readouterr().err
        assert not (run_root / "review" / "review_decision.json").exists()


def test_record_decision_cli_rejects_incompatible_mode_and_incomplete_structure(tmp_path, capsys):
    from jewelry_on_hand.cli import main

    mode_run = tmp_path / "runs" / "bad-mode"
    make_modern_analysis(
        mode_run / "analysis" / "product_analysis.json",
        product_type="手链/手串",
        detected_product_type="bracelet",
        confirmed_product_type="bracelet",
    )
    declare_scene_output_role(mode_run)
    assert main(
        [
            "record-decision",
            "--run-root",
            str(mode_run),
            "--action",
            "generate_rank_1",
            "--fidelity-confirmed",
            "--display-mode",
            "hand_held",
            "--output-role",
            "hand_worn",
        ]
    ) != 0
    assert "手串/手链与手持展示模式不兼容" in capsys.readouterr().err

    structure_run = tmp_path / "runs" / "bad-structure"
    make_modern_analysis(structure_run / "analysis" / "product_analysis.json")
    declare_scene_output_role(structure_run)
    assert main(
        [
            "record-decision",
            "--run-root",
            str(structure_run),
            "--action",
            "generate_rank_1",
            "--fidelity-confirmed",
            "--confirmed-product-type",
            "pendant_necklace",
            "--output-role",
            "hand_worn",
        ]
    ) != 0
    assert "完整主吊坠结构" in capsys.readouterr().err
    assert not (structure_run / "review" / "review_decision.json").exists()


def test_record_decision_cli_rejects_illegal_corrections_for_non_generation_actions(
    tmp_path, capsys
):
    from jewelry_on_hand.cli import main

    cases = (
        ("rerank", [], "flat_lay_source", "重新执行 prepare-review"),
        (
            "manual_reference",
            ["--manual-reference", "manual.jpg"],
            "unknown_source",
            "重新执行 prepare-review",
        ),
    )
    for index, (action, extra_args, source_type, expected) in enumerate(cases):
        run_root = tmp_path / "runs" / f"non-generation-{index}"
        analysis_path = run_root / "analysis" / "product_analysis.json"
        original = make_modern_analysis(
            analysis_path,
            product_type="普通项链",
            detected_product_type="necklace",
            confirmed_product_type="necklace",
            length_category="collarbone",
        )
        declare_scene_output_role(run_root)

        assert main(
            [
                "record-decision",
                "--run-root",
                str(run_root),
                "--action",
                action,
                *extra_args,
                "--source-image-type",
                source_type,
                "--output-role",
                "hand_worn",
            ]
        ) != 0
        assert expected in capsys.readouterr().err
        assert read_json(analysis_path) == original
        assert not (run_root / "review" / "review_decision.json").exists()


def test_record_decision_cli_rolls_back_analysis_and_decision_on_second_replace_failure(
    tmp_path, monkeypatch, capsys
):
    import os

    from jewelry_on_hand.cli import main

    run_root = tmp_path / "runs" / "rollback"
    analysis_path = run_root / "analysis" / "product_analysis.json"
    decision_path = run_root / "review" / "review_decision.json"
    original_analysis = make_modern_analysis(
        analysis_path,
        product_type="戒指",
        detected_product_type="ring",
        confirmed_product_type="ring",
        classification_evidence=["左手无名指根部可见单枚戒指"],
        wear_position="左手无名指根部",
        visible_appearance="单枚银色戒指",
        length_category=None,
        ring_count=1,
        hand_side="left",
        finger_position="ring",
        ring_wear_style="finger_base",
    )
    write_json(decision_path, {"action": "rerank", "selected_ranks": []})
    declare_scene_output_role(run_root)
    old_analysis = analysis_path.read_bytes()
    old_decision = decision_path.read_bytes()
    original_replace = os.replace
    replace_count = 0

    def fail_second_replace(source, target):
        nonlocal replace_count
        replace_count += 1
        if replace_count == 2:
            raise OSError("模拟第二次替换失败")
        return original_replace(source, target)

    monkeypatch.setattr("jewelry_on_hand.review_decision.os.replace", fail_second_replace)

    assert main(
        [
            "record-decision",
            "--run-root",
            str(run_root),
            "--action",
            "rerank",
            "--hand-side",
            "right",
            "--output-role",
            "hand_worn",
        ]
    ) != 0

    assert "文件提交失败" in capsys.readouterr().err
    assert analysis_path.read_bytes() == old_analysis
    assert decision_path.read_bytes() == old_decision


def test_record_decision_cli_imports_custom_constraints_to_canonical(tmp_path):
    from jewelry_on_hand.cli import main

    run_root = tmp_path / "runs" / "canonical-import"
    analysis_data = make_modern_analysis(
        run_root / "analysis" / "product_analysis.json",
        product_type="普通项链",
        detected_product_type="necklace",
        confirmed_product_type="necklace",
        classification_confidence="high",
        classification_evidence=["完整链条围绕颈部"],
        classification_source="auto_confirmed",
        length_category="collarbone",
        special_requirements=["保留链条连接结构"],
    )
    prepare_snapshot_decision_run(run_root)
    canonical_path = run_root / "analysis" / "product_fidelity_constraints.json"
    imported_path = run_root / "review" / "imported-constraints.json"
    imported = make_constraints(
        imported_path,
        review_status="not_applicable",
        analysis_data=analysis_data,
    )
    imported["source"]["product_id"] = "imported-source"
    write_json(imported_path, imported)
    declare_scene_output_role(run_root)

    assert main(
        [
            "record-decision",
            "--run-root",
            str(run_root),
            "--action",
            "generate_rank_1",
            "--fidelity-confirmed",
            "--fidelity-constraints-path",
            str(imported_path),
            "--output-role",
            "hand_worn",
        ]
    ) == 0

    canonical = read_json(canonical_path)
    decision = read_json(run_root / "review" / "review_decision.json")
    assert canonical["source"]["product_id"] == "imported-source"
    assert canonical["review_status"] == "not_applicable"
    assert read_json(imported_path)["review_status"] == "not_applicable"
    assert decision["fidelity_constraints_path"] == (
        "analysis/product_fidelity_constraints.json"
    )


def test_record_decision_cli_rejects_legacy_v1_necklace_without_writing(
    tmp_path,
    capsys,
) -> None:
    from jewelry_on_hand.cli import main

    run_root = tmp_path / "runs" / "legacy-v1-necklace"
    analysis_path = run_root / "analysis" / "product_analysis.json"
    analysis = make_modern_analysis(
        analysis_path,
        product_type="普通项链",
        detected_product_type="necklace",
        confirmed_product_type="necklace",
        classification_confidence="high",
        classification_evidence=["同一条双圈长链"],
        classification_source="manual_override",
        layer_count=2,
        length_category="long",
        visible_appearance="同一条海蓝宝长链绕颈形成上下双圈",
    )
    imported_path = run_root / "review" / "legacy-v1.json"
    legacy = build_product_fidelity_constraints(
        ProductAnalysis.from_dict(analysis)
    ).to_dict()
    legacy["schema_version"] = 1
    legacy.pop("pendant_semantics")
    write_json(imported_path, legacy)
    before = analysis_path.read_bytes()
    declare_scene_output_role(run_root)

    result = main(
        [
            "record-decision",
            "--run-root",
            str(run_root),
            "--action",
            "generate_rank_1",
            "--fidelity-confirmed",
            "--fidelity-constraints-path",
            str(imported_path),
            "--output-role",
            "hand_worn",
        ]
    )

    assert result != 0
    assert "历史 v1 只读" in capsys.readouterr().err
    assert analysis_path.read_bytes() == before
    assert not (run_root / "review" / "review_decision.json").exists()
    assert not (
        run_root / "analysis" / "product_fidelity_constraints.json"
    ).exists()


def test_record_decision_cli_missing_constraints_changes_nothing(tmp_path, capsys):
    from jewelry_on_hand.cli import main

    run_root = tmp_path / "runs" / "missing-constraints"
    analysis_path = run_root / "analysis" / "product_analysis.json"
    original = make_modern_analysis(
        analysis_path,
        product_type="普通项链",
        detected_product_type="necklace",
        confirmed_product_type="necklace",
        classification_confidence="high",
        classification_evidence=["完整链条围绕颈部"],
        classification_source="auto_confirmed",
        length_category="collarbone",
    )
    declare_scene_output_role(run_root)

    assert main(
        [
            "record-decision",
            "--run-root",
            str(run_root),
            "--action",
            "generate_rank_1",
            "--fidelity-confirmed",
            "--output-role",
            "hand_worn",
        ]
    ) != 0

    assert "缺少产品保真约束导入源" in capsys.readouterr().err
    assert read_json(analysis_path) == original
    assert not (run_root / "review" / "review_decision.json").exists()


def test_record_decision_cli_updates_ring_analysis_and_snapshot_atomically(tmp_path):
    from jewelry_on_hand.cli import main

    run_root = tmp_path / "runs" / "ring-correction"
    analysis_path = run_root / "analysis" / "product_analysis.json"
    original_analysis = make_modern_analysis(
        analysis_path,
        product_type="戒指",
        detected_product_type="ring",
        confirmed_product_type="ring",
        classification_confidence="high",
        classification_source="auto_confirmed",
        classification_evidence=["手指根部结构疑似戒指"],
        wear_position="左手无名指根部",
        visible_appearance="单枚银色环状首饰",
        composition="手部近景",
        occluded_parts=["戒圈背面"],
        uncertain_details=["镶嵌背面结构"],
        ring_count=1,
        hand_side="left",
        finger_position="ring",
        ring_wear_style="finger_base",
    )
    prepare_snapshot_decision_run(run_root)
    corrected_analysis = original_analysis | {
        "classification_source": "manual_override",
        "hand_side": "right",
        "finger_position": "ring",
        "ring_wear_style": "finger_base",
    }
    imported_constraints = build_product_fidelity_constraints(
        ProductAnalysis.from_dict(corrected_analysis)
    ).to_dict()
    imported_constraints_path = tmp_path / "pending-ring-constraints.json"
    write_json(imported_constraints_path, imported_constraints)
    declare_scene_output_role(run_root)

    assert main(
        [
            "record-decision",
            "--run-root",
            str(run_root),
            "--action",
            "generate_rank_1",
            "--fidelity-confirmed",
            "--fidelity-constraints-path",
            str(imported_constraints_path),
            "--confirmed-product-type",
            "ring",
            "--ring-count",
            "1",
            "--hand-side",
            "right",
            "--finger-position",
            "ring",
            "--ring-wear-style",
            "finger_base",
            "--output-role",
            "hand_worn",
        ]
    ) == 0

    analysis = read_json(analysis_path)
    decision = read_json(run_root / "review" / "review_decision.json")
    canonical = read_json(
        run_root / "analysis" / "product_fidelity_constraints.json"
    )
    assert analysis["classification_source"] == "manual_override"
    assert canonical == imported_constraints | {"review_status": "confirmed"}
    assert canonical["must_keep"]
    assert decision["fidelity_constraints_path"] == (
        "analysis/product_fidelity_constraints.json"
    )
    assert {
        field_name: analysis[field_name]
        for field_name in (
            "ring_count",
            "hand_side",
            "finger_position",
            "ring_wear_style",
        )
    } == {
        "ring_count": 1,
        "hand_side": "right",
        "finger_position": "ring",
        "ring_wear_style": "finger_base",
    }
    assert decision["confirmation_snapshot"] == {
        "confirmed_product_type": "ring",
        "source_image_type": "worn_source",
        "display_mode": "worn",
        "layer_count": 1,
        "length_category": None,
        "has_pendant": False,
        "pendant_count": 0,
        "pendant_layer": None,
        "pendant_position": None,
        "pendant_orientation": None,
        "connection_structure": None,
        "is_independent_multi_item": False,
        "ring_count": 1,
        "hand_side": "right",
        "finger_position": "ring",
        "ring_wear_style": "finger_base",
    }


def test_record_decision_cli_invalid_ring_correction_changes_nothing(tmp_path, capsys):
    from jewelry_on_hand.cli import main

    run_root = tmp_path / "runs" / "ring-invalid"
    analysis_path = run_root / "analysis" / "product_analysis.json"
    decision_path = run_root / "review" / "review_decision.json"
    original_analysis = make_modern_analysis(
        analysis_path,
        product_type="疑似戒指",
        classification_evidence=["手指根部结构疑似戒指"],
        wear_position="手指",
        composition="手部近景",
    )
    original_decision = {"action": "rerank", "selected_ranks": []}
    write_json(decision_path, original_decision)
    declare_scene_output_role(run_root)

    assert main(
        [
            "record-decision",
            "--run-root",
            str(run_root),
            "--action",
            "rerank",
            "--confirmed-product-type",
            "ring",
            "--ring-count",
            "2",
            "--hand-side",
            "left",
            "--finger-position",
            "ring",
            "--ring-wear-style",
            "finger_base",
            "--output-role",
            "hand_worn",
        ]
    ) != 0

    assert "只支持单枚戒指" in capsys.readouterr().err
    assert read_json(analysis_path) == original_analysis
    assert read_json(decision_path) == original_decision


def test_qc_cli_writes_qc_json(tmp_path):
    from jewelry_on_hand.cli import main

    generation_dir, fidelity_json, checklist_json, reference_json = (
        _ready_cli_qc_generation(tmp_path)
    )

    assert (
        main(
            [
                "qc",
                "--generation-dir",
                str(generation_dir),
                "--status",
                "rerun",
                "--passed",
                "无水印,构图正确",
                "--failed",
                "主珠被裁切",
                "--notes",
                "复跑",
                "--fidelity-checks-json",
                str(fidelity_json),
                "--checklist-checks-json",
                str(checklist_json),
                "--reference-preservation-checks-json",
                str(reference_json),
            ]
        )
        == 0
    )

    payload = read_json(generation_dir / "qc.json")
    assert payload["status"] == "rerun"
    assert payload["passed"] == ["无水印", "构图正确"]
    assert payload["failed"] == ["主珠被裁切"]
    assert payload["notes"] == "复跑"
    assert payload["reference_preservation_checks"]
    assert payload["checklist_checks"]


def test_qc_cli_writes_fidelity_checks_from_json(tmp_path):
    from jewelry_on_hand.cli import main

    generation_dir, checks_json, checklist_json, reference_json = (
        _ready_cli_qc_generation(tmp_path)
    )
    checks = read_json(checks_json)
    assert checks

    assert (
        main(
            [
                "qc",
                "--generation-dir",
                str(generation_dir),
                "--status",
                "rerun",
                "--failed",
                "关键识别点失败",
                "--fidelity-checks-json",
                str(checks_json),
                "--checklist-checks-json",
                str(checklist_json),
                "--reference-preservation-checks-json",
                str(reference_json),
            ]
        )
        == 0
    )

    assert read_json(generation_dir / "qc.json")["fidelity_checks"] == checks


def test_qc_cli_strict_critical_failures_accepts_repeated_and_csv(tmp_path):
    from jewelry_on_hand.cli import main

    generation_dir, fidelity_json, checklist_json, reference_json = (
        _ready_cli_qc_generation(tmp_path)
    )

    assert main(
        [
            "qc",
            "--generation-dir",
            str(generation_dir),
            "--status",
            "reject",
            "--fidelity-checks-json",
            str(fidelity_json),
            "--checklist-checks-json",
            str(checklist_json),
            "--reference-preservation-checks-json",
            str(reference_json),
            "--critical-failures",
            "auto_chain_added,layer_count_mismatch",
            "--critical-failures",
            "source_person_region_migrated",
        ]
    ) == 0
    assert read_json(generation_dir / "qc.json")["critical_failures"] == [
        "auto_chain_added",
        "layer_count_mismatch",
        "source_person_region_migrated",
    ]


@pytest.mark.parametrize(
    "critical_args",
    [
        ["--critical-failures"],
        ["--critical-failures", ""],
        ["--critical-failures", ","],
        ["--critical-failures", "auto_chain_added,"],
        ["--critical-failures", ",auto_chain_added"],
        ["--critical-failures", "auto_chain_added,,layer_count_mismatch"],
        ["--critical-failures", "auto_chain_added,   "],
    ],
)
def test_qc_cli_strict_critical_failures_rejects_empty_segments(
    tmp_path,
    capsys,
    critical_args,
):
    from jewelry_on_hand.cli import main

    generation_dir, fidelity_json, checklist_json, reference_json = (
        _ready_cli_qc_generation(tmp_path)
    )
    assert main(
        [
            "qc",
            "--generation-dir",
            str(generation_dir),
            "--status",
            "reject",
            "--fidelity-checks-json",
            str(fidelity_json),
            "--checklist-checks-json",
            str(checklist_json),
            "--reference-preservation-checks-json",
            str(reference_json),
            *critical_args,
        ]
    ) != 0
    assert "critical-failures 不能包含空值" in capsys.readouterr().err
    assert not (generation_dir / "qc.json").exists()


@pytest.mark.parametrize(
    ("status", "critical_args", "message"),
    [
        ("reject", ["unknown_failure"], "未知错误代码"),
        (
            "reject",
            ["auto_chain_added", "auto_chain_added,layer_count_mismatch"],
            "不能包含重复",
        ),
        ("pass", ["auto_chain_added"], "不得标记为 pass"),
    ],
)
def test_qc_cli_critical_failures_preserves_model_validation(
    tmp_path,
    capsys,
    status,
    critical_args,
    message,
):
    from jewelry_on_hand.cli import main

    generation_dir, fidelity_json, checklist_json, reference_json = (
        _ready_cli_qc_generation(tmp_path)
    )
    args = [
        "qc",
        "--generation-dir",
        str(generation_dir),
        "--status",
        status,
        "--fidelity-checks-json",
        str(fidelity_json),
        "--checklist-checks-json",
        str(checklist_json),
        "--reference-preservation-checks-json",
        str(reference_json),
    ]
    for value in critical_args:
        args.extend(["--critical-failures", value])

    assert main(args) != 0
    assert message in capsys.readouterr().err
    assert not (generation_dir / "qc.json").exists()


def test_generate_cli_builds_prompts_after_review_gate(tmp_path, monkeypatch):
    from jewelry_on_hand.cli import main

    run_root = tmp_path / "runs" / "demo"
    analysis_dir = run_root / "analysis"
    review_dir = run_root / "review"
    input_dir = run_root / "input"
    analysis_dir.mkdir(parents=True)
    review_dir.mkdir()
    input_dir.mkdir()
    product = input_dir / "product-on-hand.jpg"
    product.write_bytes(b"product")
    ref = tmp_path / "ref.jpg"
    ref.write_bytes(b"ref")
    write_json(
        analysis_dir / "product_analysis.json",
        {
            "product_type": "手链/手串",
            "wear_position": "手腕",
            "visible_appearance": "深红主珠",
            "color_family": ["深红"],
            "style_mood": "暗调闪光",
            "composition": "手腕近景",
            "product_dimensions": {},
            "needs_full_front_display": True,
            "special_requirements": ["保留主珠"],
        },
    )
    write_json(
        analysis_dir / "selected_references.json",
        [
            {
                "selected_reference": str(ref),
                "score": 80,
                "rank": 1,
                "reason": ["匹配"],
                "risk": [],
                "ignored_reference_jewelry": [],
                "metadata": {
                    "序号": 1,
                    "文件名": "ref.jpg",
                    "用途分类": "上手姿势/手模构图参考",
                    "风格分类": "暗调闪光",
                    "场景关键词": "车内 闪光",
                    "饰品类型": "手链/手串",
                    "推荐使用方式": "近景手腕",
                    "备注": "手腕/前臂露出面积足",
                    "判断置信度": "高",
                },
            }
        ],
    )
    prepare_snapshot_decision_run(run_root)
    write_review_bundle(
        RunPaths(root=run_root),
        {
            "action": "generate_rank_1",
            "selected_ranks": [1],
            "fidelity_confirmed": True,
            "output_role": "hand_worn",
        },
    )
    calls = []

    def fake_run_generation(
        paths,
        product_image,
        prompts_by_rank,
        helper_script,
        wait=True,
        **audit_inputs,
    ):
        calls.append((paths, product_image, prompts_by_rank, helper_script, wait))
        return [paths.generation_dir / "01"]

    monkeypatch.setattr("jewelry_on_hand.cli.run_generation", fake_run_generation)

    assert (
        main(
            [
                "generate",
                "--run-root",
                str(run_root),
                "--helper-script",
                str(tmp_path / "helper.py"),
                "--no-wait",
            ]
        )
        == 0
    )

    assert len(calls) == 1
    assert calls[0][1] == product
    assert list(calls[0][2]) == [1]
    assert "内部图1" in calls[0][2][1]
    assert calls[0][4] is False


def test_generate_cli_accepts_selected_reference_without_metadata(tmp_path, monkeypatch):
    from jewelry_on_hand.cli import main

    run_root = tmp_path / "runs" / "demo"
    analysis_dir = run_root / "analysis"
    review_dir = run_root / "review"
    input_dir = run_root / "input"
    analysis_dir.mkdir(parents=True)
    review_dir.mkdir()
    input_dir.mkdir()
    product = input_dir / "product-on-hand.jpg"
    product.write_bytes(b"product")
    ref = tmp_path / "ref.jpg"
    ref.write_bytes(b"ref")
    write_json(
        analysis_dir / "product_analysis.json",
        {
            "product_type": "手链/手串",
            "wear_position": "手腕",
            "visible_appearance": "深红主珠",
            "color_family": ["深红"],
            "style_mood": "暗调闪光",
            "composition": "手腕近景",
            "product_dimensions": {},
            "needs_full_front_display": True,
            "special_requirements": ["保留主珠"],
        },
    )
    write_json(
        analysis_dir / "selected_references.json",
        [
            {
                "selected_reference": str(ref),
                "score": 80,
                "rank": 2,
                "reason": ["匹配"],
                "risk": [],
                "ignored_reference_jewelry": [],
            }
        ],
    )
    prepare_snapshot_decision_run(run_root)
    paths = RunPaths(root=run_root)
    write_review_bundle(
        paths,
        {
            "action": "generate_rank_1",
            "selected_ranks": [1],
            "fidelity_confirmed": True,
            "output_role": "hand_worn",
        },
    )
    selected = read_json(paths.analysis_dir / "selected_references.json")
    selected[0].pop("metadata")
    write_json(paths.analysis_dir / "selected_references.json", selected)
    calls = []

    def fake_run_generation(paths, product_image, prompts_by_rank, helper_script, wait=True):
        calls.append((paths, product_image, prompts_by_rank, helper_script, wait))
        return [paths.generation_dir / "02"]

    monkeypatch.setattr("jewelry_on_hand.cli.run_generation", fake_run_generation)

    assert (
        main(
            [
                "generate",
                "--run-root",
                str(run_root),
                "--helper-script",
                str(tmp_path / "helper.py"),
            ]
        )
        != 0
    )

    assert calls == []


def test_generate_cli_only_builds_prompts_for_approved_selected_ranks(tmp_path, monkeypatch):
    from jewelry_on_hand.cli import main

    run_root = tmp_path / "runs" / "demo"
    analysis_dir = run_root / "analysis"
    review_dir = run_root / "review"
    input_dir = run_root / "input"
    analysis_dir.mkdir(parents=True)
    review_dir.mkdir()
    input_dir.mkdir()
    product = input_dir / "product-on-hand.jpg"
    product.write_bytes(b"product")
    references = []
    for rank in (1, 2, 3):
        ref = tmp_path / f"ref{rank}.jpg"
        ref.write_bytes(f"ref-{rank}".encode())
        references.append(
            {
                "selected_reference": str(ref),
                "score": 80 + rank,
                "rank": rank,
                "reason": [f"匹配 {rank}"],
                "risk": [],
                "ignored_reference_jewelry": [],
            }
        )
    write_json(
        analysis_dir / "product_analysis.json",
        {
            "product_type": "手链/手串",
            "wear_position": "手腕",
            "visible_appearance": "深红主珠",
            "color_family": ["深红"],
            "style_mood": "暗调闪光",
            "composition": "手腕近景",
            "product_dimensions": {},
            "needs_full_front_display": True,
            "special_requirements": ["保留主珠"],
        },
    )
    prepare_snapshot_decision_run(run_root)
    paths = RunPaths(root=run_root)
    write_review_bundle(
        paths,
        {
            "action": "generate_rank_1",
            "selected_ranks": [1],
            "fidelity_confirmed": True,
            "output_role": "hand_worn",
        },
    )
    built_ranks = []
    calls = []

    def fake_build_prompt(
        product_analysis,
        scored_reference,
        fidelity_constraints,
        output_role,
        reference_snapshot,
    ):
        assert fidelity_constraints.review_status in {"confirmed", "not_applicable"}
        assert output_role.value == "hand_worn"
        built_ranks.append(scored_reference.rank)
        return f"prompt-rank-{scored_reference.rank}"

    def fake_run_generation(
        paths,
        product_image,
        prompts_by_rank,
        helper_script,
        wait=True,
        **audit_inputs,
    ):
        calls.append((paths, product_image, prompts_by_rank, helper_script, wait))
        return [paths.generation_dir / "01"]

    monkeypatch.setattr("jewelry_on_hand.cli.build_prompt", fake_build_prompt)
    monkeypatch.setattr("jewelry_on_hand.cli.run_generation", fake_run_generation)

    assert main(["generate", "--run-root", str(run_root), "--helper-script", str(tmp_path / "helper.py")]) == 0

    assert built_ranks == [1]
    assert len(calls) == 1
    assert list(calls[0][2]) == [1]
    assert calls[0][2] == {1: "prompt-rank-1"}


def test_generate_cli_returns_nonzero_when_decision_rank_is_missing(tmp_path, monkeypatch):
    from jewelry_on_hand.cli import main

    run_root = tmp_path / "runs" / "demo"
    analysis_dir = run_root / "analysis"
    review_dir = run_root / "review"
    input_dir = run_root / "input"
    analysis_dir.mkdir(parents=True)
    review_dir.mkdir()
    input_dir.mkdir()
    (input_dir / "product-on-hand.jpg").write_bytes(b"product")
    references = []
    for rank in (1, 2):
        ref = tmp_path / f"ref{rank}.jpg"
        ref.write_bytes(f"ref-{rank}".encode())
        references.append(
            {
                "selected_reference": str(ref),
                "score": 80 + rank,
                "rank": rank,
                "reason": [f"匹配 {rank}"],
                "risk": [],
                "ignored_reference_jewelry": [],
            }
        )
    write_json(
        analysis_dir / "product_analysis.json",
        {
            "product_type": "手链/手串",
            "wear_position": "手腕",
            "visible_appearance": "深红主珠",
            "color_family": ["深红"],
            "style_mood": "暗调闪光",
            "composition": "手腕近景",
            "product_dimensions": {},
            "needs_full_front_display": True,
            "special_requirements": ["保留主珠"],
        },
    )
    write_json(analysis_dir / "selected_references.json", references)
    make_constraints(analysis_dir / "product_fidelity_constraints.json")
    declare_scene_output_role(run_root)
    write_json(review_dir / "review_decision.json", {"action": "generate_selected", "selected_ranks": [3], "fidelity_confirmed": True, "output_role": "hand_worn"})
    calls = []

    def fake_run_generation(paths, product_image, prompts_by_rank, helper_script, wait=True):
        calls.append((paths, product_image, prompts_by_rank, helper_script, wait))
        return []

    monkeypatch.setattr("jewelry_on_hand.cli.run_generation", fake_run_generation)

    assert main(["generate", "--run-root", str(run_root), "--helper-script", str(tmp_path / "helper.py")]) != 0
    assert calls == []


def test_generate_cli_reruns_supported_product_gate_before_generation(tmp_path, monkeypatch):
    from jewelry_on_hand.cli import main

    run_root = tmp_path / "runs" / "demo"
    analysis_dir = run_root / "analysis"
    review_dir = run_root / "review"
    input_dir = run_root / "input"
    analysis_dir.mkdir(parents=True)
    review_dir.mkdir()
    input_dir.mkdir()
    (input_dir / "product-on-hand.jpg").write_bytes(b"product")
    ref = tmp_path / "ref.jpg"
    ref.write_bytes(b"ref")
    analysis = make_analysis(analysis_dir / "product_analysis.json")
    analysis["product_type"] = "戒指"
    write_json(analysis_dir / "product_analysis.json", analysis)
    write_json(analysis_dir / "selected_references.json", [_selected_reference_payload(1, ref)])
    make_constraints(analysis_dir / "product_fidelity_constraints.json")
    declare_scene_output_role(run_root)
    write_json(review_dir / "review_decision.json", {"action": "generate_rank_1", "selected_ranks": [1], "fidelity_confirmed": True, "output_role": "hand_worn"})
    calls = []

    def fake_run_generation(paths, product_image, prompts_by_rank, helper_script, wait=True):
        calls.append((paths, product_image, prompts_by_rank, helper_script, wait))
        return []

    monkeypatch.setattr("jewelry_on_hand.cli.run_generation", fake_run_generation)

    assert main(["generate", "--run-root", str(run_root), "--helper-script", str(tmp_path / "helper.py")]) != 0
    assert calls == []


def test_generate_cli_preserves_mirror_relative_path_from_review_package(tmp_path, monkeypatch):
    from jewelry_on_hand.cli import main

    paths_root = tmp_path / "runs"
    run_root = paths_root / "demo"
    product = run_root / "input" / "product-on-hand.jpg"
    product.parent.mkdir(parents=True)
    product.write_bytes(b"product")
    paths = RunPaths(root=run_root)
    paths.analysis_dir.mkdir(parents=True, exist_ok=True)
    paths.review_dir.mkdir(parents=True, exist_ok=True)
    make_analysis(paths.analysis_dir / "product_analysis.json")

    original_ref = tmp_path / "plain.jpg"
    original_ref.write_bytes(b"ref")
    scored = ScoredReference(
        ReferenceRow(
            index=1,
            file_name="plain.jpg",
            relative_path="reference/对镜/plain.jpg",
            absolute_path=original_ref,
            width=100,
            height=200,
            size_mb=0.1,
            purpose_category="上手姿势/手模构图参考",
            bracelet_applicability="是",
            default_strategy="常规可优先使用",
            style_category="暗调闪光",
            scene_keywords="车内",
            jewelry_type="手链/手串",
            recommended_usage="近景",
            notes="正面视角，主体居中，手腕露出，无文字或 UI",
            confidence="高",
            file_exists=True,
            framing="手部近景",
            visible_body_regions="左手腕 / 前臂完整露出",
            product_visibility="展示面积充足，大于 35%",
            collar_type="无衣领",
            clothing_occlusion_risk="衣物无遮挡",
            pose_keywords="身体未入镜，前臂自然抬起",
            existing_jewelry="左手腕原有手链",
            crop_risk="裁切风险低",
            hand_side="左手",
            hand_orientation="手背朝向镜头",
        ),
        score=99,
        rank=1,
        reason=["匹配"],
        risk=[],
        ignored_reference_jewelry=[],
    )
    analysis = load_product_analysis(paths.analysis_dir / "product_analysis.json")
    snapshot = build_candidate_snapshot(analysis, scored, "hand_worn")
    write_review_package(
        paths,
        product,
        [scored],
        [scored],
        composition_snapshots=[snapshot],
    )
    constraints = build_product_fidelity_constraints(analysis).to_dict()
    if constraints["review_status"] == "pending":
        constraints["review_status"] = "confirmed"
    write_json(paths.analysis_dir / "product_fidelity_constraints.json", constraints)
    declare_scene_output_role(run_root)
    write_review_bundle(
        paths,
        {
            "action": "generate_rank_1",
            "selected_ranks": [1],
            "fidelity_confirmed": True,
            "output_role": "hand_worn",
        },
    )
    calls = []

    def fake_run_generation(
        paths,
        product_image,
        prompts_by_rank,
        helper_script,
        wait=True,
        **audit_inputs,
    ):
        calls.append((paths, product_image, prompts_by_rank, helper_script, wait))
        return [paths.generation_dir / "01"]

    monkeypatch.setattr("jewelry_on_hand.cli.run_generation", fake_run_generation)

    assert main(["generate", "--run-root", str(run_root), "--helper-script", str(tmp_path / "helper.py")]) == 0

    assert len(calls) == 1
    assert "前景手部 + 镜中反射手部" not in calls[0][2][1]
    assert "内部图1" in calls[0][2][1]


def test_rerank_batch_cli_rewrites_review_packages_without_reusing_files(tmp_path):
    from jewelry_on_hand.cli import main

    output_root = tmp_path / "runs"
    shared_references = [tmp_path / f"ref-{index}.jpg" for index in range(1, 7)]
    for reference in shared_references:
        reference.write_bytes(reference.name.encode("utf-8"))

    run_roots = []
    for run_name in ("sku-a", "sku-b"):
        run_root = output_root / run_name
        paths = RunPaths(root=run_root)
        paths.input_dir.mkdir(parents=True)
        paths.analysis_dir.mkdir()
        paths.review_dir.mkdir()
        (paths.review_dir / "rank-9-stale.jpg").write_bytes(b"stale")
        product = paths.input_dir / "product-on-hand.jpg"
        product.write_bytes(b"product")
        make_analysis(paths.analysis_dir / "product_analysis.json")
        declare_scene_output_role(run_root)
        candidates = [
            _candidate_payload(index, reference, score=101 - index)
            for index, reference in enumerate(shared_references, start=1)
        ]
        write_json(paths.analysis_dir / "reference_candidates.json", candidates)
        run_roots.append(run_root)

    assert main([
        "rerank-batch",
        "--output-root",
        str(output_root),
        "--run-id",
        "sku-a",
        "--run-id",
        "sku-b",
    ]) == 0

    selected_files = []
    for run_root in run_roots:
        selected = read_json(run_root / "analysis" / "selected_references.json")
        assert [item["rank"] for item in selected] == [1, 2, 3]
        assert (run_root / "review" / "review.html").is_file()
        assert len(list((run_root / "review").glob("rank-*"))) == 3
        selected_files.extend(item["metadata"]["source_file_name"] for item in selected)

    assert len(selected_files) == len(set(selected_files)) == 6

def _selected_reference_payload(rank, path):
    return {
        "selected_reference": str(path),
        "score": 80,
        "rank": rank,
        "reason": ["匹配"],
        "risk": [],
        "ignored_reference_jewelry": [],
        "metadata": {},
    }


def _candidate_payload(index, path, score):
    return {
        "selected_reference": str(path),
        "score": score,
        "rank": index,
        "reason": ["匹配"],
        "risk": [],
        "ignored_reference_jewelry": [],
        "metadata": {
            "index": index,
            "file_name": path.name,
            "relative_path": path.name,
            "absolute_path": str(path),
            "width": 100,
            "height": 200,
            "size_mb": 0.1,
            "用途分类": "手部佩戴图",
            "purpose_category": "手部佩戴图",
            "手链手串适用性": "是",
            "bracelet_applicability": "是",
            "默认使用策略": "常规可优先使用",
            "default_strategy": "常规可优先使用",
            "风格分类": f"风格-{index}",
            "style_category": f"风格-{index}",
            "场景关键词": f"场景-{index}",
            "scene_keywords": f"场景-{index}",
            "饰品类型": "手链/手串",
            "jewelry_type": "手链/手串",
            "推荐使用方式": f"姿势-{index}",
            "recommended_usage": f"姿势-{index}",
            "备注": "正面视角，主体居中，手腕完整，无文字或 UI",
            "notes": "正面视角，主体居中，手腕完整，无文字或 UI",
            "判断置信度": "高",
            "confidence": "高",
            "file_exists": True,
            "framing": "手部近景",
            "visible_body_regions": "左手腕 / 前臂完整露出",
            "product_visibility": "展示面积充足，大于 35%",
            "collar_type": "无衣领",
            "clothing_occlusion_risk": "衣物无遮挡",
            "pose_keywords": "身体未入镜，前臂自然抬起",
            "existing_jewelry": "左手腕原有手链",
            "crop_risk": "裁切风险低",
            "hand_side": "左手",
            "hand_orientation": "手背朝向镜头",
        },
    }


def test_prepare_review_rejects_ignore_pending_enrichment_with_local_classification_before_run_creation(
    tmp_path,
    monkeypatch,
    capsys,
):
    from jewelry_on_hand.cli import main

    product = tmp_path / "product.jpg"
    product.write_bytes(b"product")
    analysis = tmp_path / "analysis.json"
    make_analysis(analysis)
    reference = tmp_path / "reference.jpg"
    reference.write_bytes(b"reference")
    classification = tmp_path / "classification.xlsx"
    make_catalog(classification, reference)
    output_root = tmp_path / "runs"

    def fail_if_called(*_args, **_kwargs):
        raise AssertionError("互斥校验必须发生在线上同步前")

    monkeypatch.setattr(
        "jewelry_on_hand.cli.sync_and_load_reference_rows",
        fail_if_called,
    )

    result = main(
        [
            "prepare-review",
            "--product-image",
            str(product),
            "--analysis-json",
            str(analysis),
            "--classification",
            str(classification),
            "--ignore-pending-enrichment",
            "--output-root",
            str(output_root),
            "--run-id",
            "conflict",
            "--output-role",
            "hand_worn",
        ]
    )

    assert result == 1
    assert "不能与 --classification 同时使用" in capsys.readouterr().err
    assert not (output_root / "conflict").exists()


def test_prepare_review_with_pending_ignore_writes_reference_source_snapshot(
    tmp_path,
    monkeypatch,
):
    import hashlib

    from jewelry_on_hand.cli import main
    from jewelry_on_hand.reference_catalog import load_reference_rows

    product = tmp_path / "product.jpg"
    product.write_bytes(b"product")
    analysis = tmp_path / "analysis.json"
    make_analysis(analysis)
    reference = tmp_path / "reference.jpg"
    reference.write_bytes(b"reference")
    catalog = tmp_path / "catalog.xlsx"
    make_catalog(catalog, reference)
    cache_root = tmp_path / "reference-cache"
    calls = []

    def fake_sync(config, *, ignore_pending_enrichment=False):
        calls.append(ignore_pending_enrichment)
        write_json(
            config.cache_root / "manifest.json",
            {
                "source": {
                    "wiki_url": "https://my.feishu.cn/wiki/example",
                    "table_name": "素材收录池",
                    "base_token": "base-token",
                    "table_id": "table-id",
                },
                "records": [
                    {
                        "record_id": "rec1",
                        "stable_index": 1,
                        "source_fields": {"素材编号": "RP000001"},
                        "pending_enrichment": False,
                        "usable": True,
                    },
                    {
                        "record_id": "rec2",
                        "stable_index": 2,
                        "source_fields": {"素材编号": "RP000308"},
                        "pending_enrichment": True,
                        "usable": True,
                    },
                    {
                        "record_id": "rec3",
                        "stable_index": 3,
                        "source_fields": {"素材编号": "RP000309"},
                        "pending_enrichment": True,
                        "usable": True,
                    },
                ],
            },
        )
        return load_reference_rows(catalog)

    monkeypatch.setattr(
        "jewelry_on_hand.cli.sync_and_load_reference_rows",
        fake_sync,
    )
    output_root = tmp_path / "runs"

    assert main(
        [
            "prepare-review",
            "--product-image",
            str(product),
            "--analysis-json",
            str(analysis),
            "--ignore-pending-enrichment",
            "--reference-cache-root",
            str(cache_root),
            "--output-root",
            str(output_root),
            "--run-id",
            "online",
            "--output-role",
            "hand_worn",
        ]
    ) == 0

    manifest_bytes = (cache_root / "manifest.json").read_bytes()
    snapshot = read_json(
        output_root / "online" / "analysis" / "reference_source_snapshot.json"
    )
    assert calls == [True]
    assert snapshot == {
        "schema_version": 1,
        "source": {
            "wiki_url": "https://my.feishu.cn/wiki/example",
            "table_name": "素材收录池",
            "base_token": "base-token",
            "table_id": "table-id",
        },
        "pagination_complete": True,
        "synced_total_count": 3,
        "ignored_pending_count": 2,
        "retained_usable_count": 1,
        "ignored_pending_records": [
            {"record_id": "rec2", "material_number": "RP000308"},
            {"record_id": "rec3", "material_number": "RP000309"},
        ],
        "manifest_sha256": hashlib.sha256(manifest_bytes).hexdigest(),
    }


def _task9_reference_row(tmp_path, product_type, display_mode):
    reference = tmp_path / f"{product_type}-{display_mode}-reference.jpg"
    reference.write_bytes(b"reference")
    hand_held = display_mode == "hand_held"
    return ReferenceRow(
        index=1,
        file_name=reference.name,
        relative_path=reference.name,
        absolute_path=reference,
        width=1200,
        height=1600,
        size_mb=0.2,
        purpose_category=(
            "手部佩戴图；深色背景；手持展示构图参考"
            if hand_held
            else "生活场景图；深色背景；真人佩戴构图参考"
        ),
        bracelet_applicability="否",
        default_strategy="优先使用",
        style_category="自然精致",
        scene_keywords="自然光",
        jewelry_type="项链",
        recommended_usage=(
            "锁骨佩戴展示" if not hand_held else "双手捏持，完整链条自然垂落"
        ),
        notes="无原有首饰，画面空间充足，真实接触",
        confidence="高",
        file_exists=True,
        applicable_product_types=product_type,
        applicable_display_modes=display_mode,
        framing="锁骨与胸前近景" if not hand_held else "双手与胸前近景",
        visible_body_regions="颈部、锁骨、胸前" if not hand_held else "双手、手指、掌心",
        product_visibility="高",
        neck_visibility="高" if not hand_held else "低",
        collarbone_visibility="高" if not hand_held else "低",
        chest_visibility="高",
        hand_visibility="高" if hand_held else "低",
        collar_type="低领",
        clothing_occlusion_risk="低",
        hair_occlusion_risk="低",
        pose_keywords="正面" if not hand_held else "双手捏持，链条完整",
        existing_jewelry="无",
        crop_risk="低",
    )


def _task9_ring_reference_rows(tmp_path):
    rows = []
    for index in range(1, 4):
        reference = tmp_path / f"ring-worn-reference-{index}.jpg"
        reference.write_bytes(f"ring-reference-{index}".encode())
        rows.append(
            ReferenceRow(
                index=index,
                file_name=reference.name,
                relative_path=reference.name,
                absolute_path=reference,
                width=1200,
                height=1600,
                size_mb=0.2,
                purpose_category="手部佩戴图；深色背景；戒指真人佩戴构图参考",
                bracelet_applicability="否",
                default_strategy="优先使用",
                style_category="自然手部近景",
                scene_keywords=f"深色背景 自然光 场景{index}",
                jewelry_type="戒指",
                recommended_usage="左手无名指佩戴",
                notes="目标手指完整，戒面清晰",
                confidence="高",
                file_exists=True,
                applicable_product_types="ring",
                applicable_display_modes="worn",
                framing="手部近景",
                visible_body_regions="完整手部和五指",
                product_visibility="高",
                hand_visibility="高",
                pose_keywords="手背朝上",
                existing_jewelry="原有戒指",
                crop_risk="低",
                hand_side="left",
                visible_fingers="index,middle,ring,little",
                hand_orientation="手背朝上",
                ring_face_visibility="高",
                finger_separation="高",
                finger_occlusion_risk="低",
            )
        )
    return rows


def _task9_local_helper(tmp_path):
    helper_log = tmp_path / "helper-calls.jsonl"
    helper_script = tmp_path / "local-helper.py"
    helper_script.write_text(
        "\n".join(
            [
                "import json",
                "import sys",
                "from pathlib import Path",
                "log_path = Path(__file__).with_name('helper-calls.jsonl')",
                "with log_path.open('a', encoding='utf-8') as handle:",
                "    handle.write(json.dumps(sys.argv[1:], ensure_ascii=False) + '\\n')",
                "print(json.dumps({'ok': True, 'data': {'status': 'pending', 'out_task_id': 'local-task'}}))",
            ]
        ),
        encoding="utf-8",
    )
    return helper_script, helper_log


def _task9_runtime_checklist(run_root, destination):
    analysis = ProductAnalysis.from_dict(
        read_json(run_root / "analysis" / "product_analysis.json")
    )
    constraints = ProductFidelityConstraints.from_dict(
        read_json(run_root / "analysis" / "product_fidelity_constraints.json")
    )
    checklist_context = (
        {
            "product_analysis": analysis,
            "fidelity_constraints": constraints,
        }
        if constraints.schema_version == 2
        else {}
    )
    checks = [
        {
            "id": qc_check_id(question),
            "question": question,
            "result": "pass",
            "notes": "已逐项核对",
        }
        for question in build_qc_checklist(
            analysis.normalized_product_type,
            analysis.display_mode,
            constraints.must_keep,
            **checklist_context,
        )
    ]
    write_json(destination, checks)
    return destination


def _assert_task9_submit_call(
    helper_log,
    run_root,
    expected_reference,
    generation_dir,
    expected_product=None,
):
    calls = [
        json.loads(line)
        for line in helper_log.read_text(encoding="utf-8").splitlines()
    ]
    assert len(calls) == 1
    command = calls[0]
    assert command[0] == "submit"
    assert command[command.index("--model") + 1] == "gpt_image_2"
    assert command[command.index("--aspect-ratio") + 1] == "3:4"
    assert command[command.index("--resolution") + 1] == "2K"
    assert command.count("--image") == 2
    image_indexes = [index for index, value in enumerate(command) if value == "--image"]
    expected_reference_path = Path(expected_reference)
    assert command[image_indexes[0] + 1] == str(expected_reference_path)
    reference_copy = generation_dir / (
        f"hand-reference{expected_reference_path.suffix or '.jpg'}"
    )
    assert reference_copy.read_bytes() == expected_reference_path.read_bytes()
    expected_product_path = expected_product or run_root / "input" / "product-on-hand.jpg"
    assert command[image_indexes[1] + 1] == str(expected_product_path)


def _ready_cli_qc_generation(tmp_path):
    generation_dir = tmp_path / "run" / "generation" / "01"
    generation_dir.mkdir(parents=True)
    analysis = ProductAnalysis.from_dict(
        {
            "product_type": "bracelet",
            "detected_product_type": "bracelet",
            "confirmed_product_type": "bracelet",
            "classification_confidence": "high",
            "classification_evidence": ["手腕处可见闭合珠串"],
            "classification_source": "manual_override",
            "display_mode": "worn",
            "source_image_type": "worn_source",
            "wear_position": "手腕",
            "visible_appearance": "圆珠手链主珠右侧有一颗透明随形",
            "color_family": ["海蓝", "透明"],
            "style_mood": "清透",
            "composition": "真人手腕近景",
            "product_dimensions": {"bead_diameter_mm": 8.0},
            "needs_full_front_display": True,
            "special_requirements": ["保持可见珠序"],
            "layer_count": 1,
            "has_pendant": False,
            "pendant_count": 0,
            "pendant_layer": None,
            "is_independent_multi_item": False,
        }
    )
    constraints_data = build_product_fidelity_constraints(analysis).to_dict()
    if constraints_data["review_status"] == "pending":
        constraints_data["review_status"] = "confirmed"
    constraints = ProductFidelityConstraints.from_dict(constraints_data)
    snapshot = ReferenceCompositionSnapshot(
        rank=1,
        reference_file="rank-1-scene.jpg",
        reference_sha256="1" * 64,
        output_role="hand_worn",
        framing="手腕近景",
        camera_angle="平视",
        subject_placement="手腕居中",
        visible_body_regions=("左手腕",),
        pose=ReferencePose("身体未入镜", "前臂横向", "手背朝上", "左手"),
        clothing="黑色袖口",
        background="深色木纹",
        lighting="左侧柔光",
        replacement_target=ReplacementTarget("左手腕", "原手串", 1),
        other_jewelry_to_remove=(),
        text_or_ui_risk="none",
        product_visibility_sufficient=True,
        composition_signature="signature",
    )
    write_json(
        generation_dir / "product-analysis.json",
        product_analysis_to_dict(analysis),
    )
    write_json(
        generation_dir / "product-fidelity-constraints.json",
        constraints_data,
    )
    write_json(
        generation_dir / "reference-composition-snapshot.json",
        snapshot.to_dict(),
    )
    write_json(
        generation_dir / "input-manifest.json",
        {"schema_version": 1, "output_role": "hand_worn"},
    )
    (generation_dir / "scene-reference.jpg").write_bytes(b"scene")
    (generation_dir / "product-reference.jpg").write_bytes(b"product")
    (generation_dir / "result.png").write_bytes(b"result")
    write_qc_review_page(generation_dir)

    fidelity_path = tmp_path / "fidelity-checks.json"
    write_json(
        fidelity_path,
        [
            {
                "name": item.name,
                "question": item.qc_question,
                "result": "pass",
                "notes": f"对照产品图确认 {item.name} 结构和位置一致",
            }
            for item in constraints.must_keep
        ],
    )
    checklist_path = tmp_path / "checklist-checks.json"
    write_json(
        checklist_path,
        [
            {
                "id": item.id,
                "question": item.question,
                "result": "pass",
                "notes": f"逐项检查确认：{item.question}",
            }
            for item in build_qc_checklist(
                product_analysis=analysis,
                fidelity_constraints=constraints,
            )
        ],
    )
    reference_path = tmp_path / "reference-preservation-checks.json"
    write_json(
        reference_path,
        [
            {
                "name": name,
                "question": question,
                "result": "pass",
                "notes": f"对照参考底图网格确认 {name} 保持一致",
            }
            for name, question in build_reference_preservation_checklist(snapshot)
        ],
    )
    return generation_dir, fidelity_path, checklist_path, reference_path
