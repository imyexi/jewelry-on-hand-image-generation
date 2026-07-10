from pathlib import Path

import pytest
from openpyxl import Workbook

from jewelry_on_hand import reference_catalog
from jewelry_on_hand.models import ReferenceRow
from jewelry_on_hand.reference_catalog import REQUIRED_COLUMNS, load_reference_rows


EXPECTED_COLUMNS = (
    "序号",
    "文件名",
    "相对路径",
    "绝对路径",
    "宽度",
    "高度",
    "大小MB",
    "用途分类",
    "手链手串适用性",
    "默认使用策略",
    "风格分类",
    "场景关键词",
    "饰品类型",
    "推荐使用方式",
    "备注",
    "判断置信度",
)

GENERIC_COLUMNS = (
    "适用产品类型",
    "适用展示模式",
    "人物取景范围",
    "可见身体区域",
    "产品预计展示面积",
    "颈部可见度",
    "锁骨可见度",
    "胸前可见度",
    "手部可见度",
    "衣领类型",
    "衣物遮挡风险",
    "头发遮挡风险",
    "姿势关键词",
    "镜面关系",
    "原有首饰类型",
    "裁切风险",
)


def _row(absolute_path, **overrides):
    data = {
        "序号": 1,
        "文件名": "ref.jpg",
        "相对路径": "reference/ref.jpg",
        "绝对路径": absolute_path,
        "宽度": 100,
        "高度": 200,
        "大小MB": 0.1,
        "用途分类": "上手姿势/手模构图参考",
        "手链手串适用性": "是：手腕露出",
        "默认使用策略": "常规可优先使用",
        "风格分类": "暗调闪光",
        "场景关键词": "车内 闪光",
        "饰品类型": "手链/手串",
        "推荐使用方式": "近景手腕",
        "备注": "手腕/前臂露出面积足",
        "判断置信度": "高",
    }
    data.update(overrides)
    return [data[column] for column in EXPECTED_COLUMNS]


def _generic_row(absolute_path, **overrides):
    data = dict(zip(EXPECTED_COLUMNS, _row(absolute_path), strict=True))
    data.update(
        {
            "适用产品类型": "necklace,pendant_necklace",
            "适用展示模式": "worn,hand_held",
            "人物取景范围": "胸前半身",
            "可见身体区域": "颈部 锁骨 胸前 手部",
            "产品预计展示面积": "高",
            "颈部可见度": "高",
            "锁骨可见度": "高",
            "胸前可见度": "高",
            "手部可见度": "中",
            "衣领类型": "低领",
            "衣物遮挡风险": "低",
            "头发遮挡风险": "低",
            "姿势关键词": "正面站立 单手持链",
            "镜面关系": "无镜面",
            "原有首饰类型": "细项链",
            "裁切风险": "低",
        }
    )
    data.update(overrides)
    return [data[column] for column in EXPECTED_COLUMNS + GENERIC_COLUMNS]


def _save_workbook(path, *, title="分类明细", columns=EXPECTED_COLUMNS, rows=()):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws.append(list(columns))
    for row in rows:
        ws.append(row)
    wb.save(path)


class _FakeWorksheet:
    def __init__(self, rows):
        self.rows = rows

    def iter_rows(self, values_only=True):
        assert values_only is True
        return iter(self.rows)


class _FakeWorkbook:
    def __init__(self, sheetnames, rows):
        self.sheetnames = sheetnames
        self.worksheet = _FakeWorksheet(rows)
        self.closed = False

    def __getitem__(self, name):
        assert name == "分类明细"
        return self.worksheet

    def close(self):
        self.closed = True


def test_required_columns_are_exact_chinese_headers():
    assert REQUIRED_COLUMNS == EXPECTED_COLUMNS


def test_load_reference_rows_maps_all_key_fields(tmp_path):
    image = tmp_path / "ref.jpg"
    image.write_bytes(b"fake")
    workbook = tmp_path / "catalog.xlsx"
    _save_workbook(workbook, rows=[_row(str(image))])

    rows = load_reference_rows(workbook)

    assert len(rows) == 1
    assert rows[0].index == 1
    assert rows[0].file_name == "ref.jpg"
    assert rows[0].relative_path == "reference/ref.jpg"
    assert rows[0].absolute_path == Path(image)
    assert isinstance(rows[0].absolute_path, Path)
    assert rows[0].width == 100
    assert rows[0].height == 200
    assert rows[0].size_mb == 0.1
    assert rows[0].purpose_category == "上手姿势/手模构图参考"
    assert rows[0].bracelet_applicability == "是：手腕露出"
    assert rows[0].default_strategy == "常规可优先使用"
    assert rows[0].style_category == "暗调闪光"
    assert rows[0].scene_keywords == "车内 闪光"
    assert rows[0].jewelry_type == "手链/手串"
    assert rows[0].recommended_usage == "近景手腕"
    assert rows[0].notes == "手腕/前臂露出面积足"
    assert rows[0].confidence == "高"
    assert rows[0].file_exists is True


def test_load_reference_rows_maps_optional_generic_columns(tmp_path):
    image = tmp_path / "necklace.jpg"
    image.write_bytes(b"fake")
    workbook = tmp_path / "generic-catalog.xlsx"
    _save_workbook(
        workbook,
        columns=EXPECTED_COLUMNS + GENERIC_COLUMNS,
        rows=[_generic_row(str(image))],
    )

    row = load_reference_rows(workbook)[0]

    assert row.applicable_product_types == "necklace,pendant_necklace"
    assert row.applicable_display_modes == "worn,hand_held"
    assert row.framing == "胸前半身"
    assert row.visible_body_regions == "颈部 锁骨 胸前 手部"
    assert row.product_visibility == "高"
    assert row.neck_visibility == "高"
    assert row.collarbone_visibility == "高"
    assert row.chest_visibility == "高"
    assert row.hand_visibility == "中"
    assert row.collar_type == "低领"
    assert row.clothing_occlusion_risk == "低"
    assert row.hair_occlusion_risk == "低"
    assert row.pose_keywords == "正面站立 单手持链"
    assert row.mirror_relation == "无镜面"
    assert row.existing_jewelry == "细项链"
    assert row.crop_risk == "低"


def test_legacy_workbook_keeps_generic_fields_empty_instead_of_inferring_necklace(tmp_path):
    image = tmp_path / "legacy.jpg"
    image.write_bytes(b"fake")
    workbook = tmp_path / "legacy-catalog.xlsx"
    _save_workbook(
        workbook,
        rows=[_row(str(image), 饰品类型="普通项链")],
    )

    row = load_reference_rows(workbook)[0]

    assert row.jewelry_type == "普通项链"
    assert row.applicable_product_types == ""
    assert row.applicable_display_modes == ""
    assert row.neck_visibility == ""


def test_legacy_reference_row_positional_arguments_remain_compatible(tmp_path):
    image = tmp_path / "legacy.jpg"
    image.write_bytes(b"fake")

    row = ReferenceRow(
        1,
        image.name,
        "reference/legacy.jpg",
        image,
        100,
        200,
        0.1,
        "上手姿势参考",
        "是：手腕露出",
        "常规可优先使用",
        "清透自然光",
        "手腕",
        "手链/手串",
        "近景手腕",
        "",
        "高",
        True,
    )

    assert row.bracelet_applicability == "是：手腕露出"
    assert row.applicable_product_types == ""
    assert row.pose_keywords == ""
    assert row.mirror_relation == ""


def test_missing_sheet_raises_clear_value_error(tmp_path):
    workbook = tmp_path / "catalog.xlsx"
    _save_workbook(workbook, title="其他表", rows=[])

    with pytest.raises(ValueError, match="分类明细"):
        load_reference_rows(workbook)


def test_missing_required_column_raises_clear_value_error(tmp_path):
    workbook = tmp_path / "catalog.xlsx"
    columns = tuple(column for column in EXPECTED_COLUMNS if column != "默认使用策略")
    _save_workbook(workbook, columns=columns, rows=[])

    with pytest.raises(ValueError, match="分类明细.*默认使用策略"):
        load_reference_rows(workbook)


def test_blank_rows_are_ignored(tmp_path):
    image = tmp_path / "ref.jpg"
    image.write_bytes(b"fake")
    workbook = tmp_path / "catalog.xlsx"
    _save_workbook(workbook, rows=[[None] * len(EXPECTED_COLUMNS), _row(str(image)), ["  "] * len(EXPECTED_COLUMNS)])

    rows = load_reference_rows(workbook)

    assert len(rows) == 1
    assert rows[0].file_name == "ref.jpg"


def test_missing_file_path_returns_file_exists_false(tmp_path):
    missing_image = tmp_path / "missing.jpg"
    workbook = tmp_path / "catalog.xlsx"
    _save_workbook(workbook, rows=[_row(str(missing_image))])

    rows = load_reference_rows(workbook)

    assert rows[0].absolute_path == missing_image
    assert rows[0].file_exists is False


def test_empty_optional_cells_use_none_or_empty_string(tmp_path):
    image = tmp_path / "ref.jpg"
    image.write_bytes(b"fake")
    workbook = tmp_path / "catalog.xlsx"
    _save_workbook(
        workbook,
        rows=[
            _row(
                str(image),
                宽度=None,
                高度=None,
                大小MB=None,
                备注=None,
            )
        ],
    )

    rows = load_reference_rows(workbook)

    assert rows[0].width is None
    assert rows[0].height is None
    assert rows[0].size_mb is None
    assert rows[0].notes == ""


def test_empty_required_absolute_path_raises_row_and_column_value_error(tmp_path):
    workbook = tmp_path / "catalog.xlsx"
    _save_workbook(workbook, rows=[_row(None)])

    with pytest.raises(ValueError, match="分类明细.*第 2 行.*绝对路径"):
        load_reference_rows(workbook)


@pytest.mark.parametrize(
    ("sheetnames", "rows", "raises"),
    [
        (["分类明细"], [EXPECTED_COLUMNS, _row("C:/missing/ref.jpg")], False),
        ([], [EXPECTED_COLUMNS], True),
        (["分类明细"], [EXPECTED_COLUMNS[:-1]], True),
        (["分类明细"], [EXPECTED_COLUMNS, _row(None)], True),
    ],
)
def test_workbook_is_closed_for_success_and_error_paths(monkeypatch, sheetnames, rows, raises):
    fake_workbook = _FakeWorkbook(sheetnames, rows)
    monkeypatch.setattr(reference_catalog, "load_workbook", lambda *args, **kwargs: fake_workbook)

    if raises:
        with pytest.raises(ValueError):
            reference_catalog.load_reference_rows("catalog.xlsx")
    else:
        reference_catalog.load_reference_rows("catalog.xlsx")

    assert fake_workbook.closed is True


def test_reference_row_reads_multi_category_fields(tmp_path):
    path = tmp_path / "necklace.jpg"
    path.write_bytes(b"x")
    row = ReferenceRow.from_dict(
        {
            "序号": 9,
            "文件名": path.name,
            "绝对路径": str(path),
            "适用产品类型": "普通项链,带链吊坠",
            "适用展示模式": "worn,hand_held",
            "人物取景范围": "胸前半身",
            "可见身体区域": "颈部 锁骨 胸前 手部",
            "产品预计展示面积": "高",
            "颈部可见度": "高",
            "锁骨可见度": "高",
            "胸前可见度": "高",
            "手部可见度": "中",
            "衣领类型": "低领",
            "衣物遮挡风险": "低",
            "头发遮挡风险": "低",
            "姿势关键词": "正面站立",
            "镜面关系": "镜中自拍",
            "原有首饰类型": "细项链",
            "裁切风险": "低",
            "文件存在": True,
        }
    )

    assert row.applicable_product_types == "普通项链,带链吊坠"
    assert row.applicable_display_modes == "worn,hand_held"
    assert row.framing == "胸前半身"
    assert row.product_visibility == "高"
    assert row.neck_visibility == "高"
    assert row.pose_keywords == "正面站立"
    assert row.mirror_relation == "镜中自拍"
    assert row.existing_jewelry == "细项链"
    assert "颈部" in row.combined_text()
    metadata = row.metadata_dict()
    assert metadata["适用产品类型"] == "普通项链,带链吊坠"
    assert metadata["姿势关键词"] == "正面站立"
    assert metadata["镜面关系"] == "镜中自拍"


def test_reference_row_accepts_existing_applicable_category_header(tmp_path):
    path = tmp_path / "necklace.jpg"
    path.write_bytes(b"x")

    row = ReferenceRow.from_dict(
        {
            "序号": 10,
            "文件名": path.name,
            "绝对路径": str(path),
            "适用品类": "普通项链、带链吊坠",
            "文件存在": True,
        }
    )

    assert row.applicable_product_types == "普通项链、带链吊坠"
